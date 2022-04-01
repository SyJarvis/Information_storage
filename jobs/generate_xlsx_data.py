# -*- coding: UTF-8 -*-
# @Time     : 4/1/22 12:18 AM
# @Author   : Runke Zhong
# @Software : PyCharm


import openpyxl
import random, datetime
import calendar

"""
获取当月的天数
"""
def get_month_days(year=None, month=None):

    t_date = None
    if year is None and month is None:
        date = datetime.datetime.now().strftime("%Y-%m")
        date = date.split("-")
        for i in range(len(date)):
            date[i] = int(date[i])
        t_date = tuple(date)
    else:
        t_date = tuple(year, month)

    days = calendar.monthrange(t_date[0], t_date[1])
    print(days, type(days))
    return days[1]


"""
随机生成体温
36.0-36.9
"""


def rand_int(temp):
    d = random.random()
    d = str(d)[2:3]
    return str(temp) + "." + d


"""
指定开始日期，获取之后的几天的日期
"""


def get_dates(month, start_day, count_day):

    days = get_month_days()
    dates = []
    day = start_day
    n = 1
    for i in range(count_day):

        if day < days:
            date = str(month) + "月" + str(day) + "日"
            day += 1
            dates.append(date)
        else:
            date = str(month + 1) + "月" + str(n) + "日"
            n += 1
            dates.append(date)
    return dates


def get_excel_data(filename, month, start_day, count_day, temp):
    """
    :param filename: 文件名
    :param start_day: 开始日期
    :param count_day: 多少天
    :param temp: 体温值
    :return:
    """

    workbook = openpyxl.load_workbook(filename)
    # 获取一共要生成几天的sheet表
    dates = get_dates(month, start_day, count_day)
    sheet = workbook.worksheets[0]
    # sheet = workbook[sheet_name]

    dates_datas = {}

    for i in range(len(dates)):
        datas = []
        for r in range(1, sheet.max_row):
            row_datas = []
            for j in range(1, sheet.max_column + 1):
                msg = sheet.cell(r, j).value
                if msg is None:
                    msg = ''
                row_datas.append(msg)

            print(row_datas)
            if r == 1:
                row_datas[0] = row_datas[0][:-6] + dates[i]
                print(row_datas)
            elif r >= 5:
                row_datas[4] = rand_int(temp)
                row_datas[4+2] = rand_int(temp)
                row_datas[4+2+2] = rand_int(temp)
            datas.append(row_datas)
        dates_datas[dates[i]] = datas

    # print(dates_datas)
    return dates_datas








def write_excel_data(filename, dates):

    workbook = openpyxl.Workbook()
    n=0
    for date, datas in dates.items():
        sheet = workbook.create_sheet(date, index=n)
        for r in range(1, len(datas)+1):
            for c in range(1, len(datas[r-1])+1):
                data = datas[r-1][c-1]
                sheet.cell(r, c, data)
                print("sheet", sheet.cell(r, c).value)

        n+=1

    try:
        workbook.save(filename)
    except Exception as e:
        print(e)
    print("ok")
    return filename



dates = get_excel_data("./jobs/21软件7班 早午晚检3.xlsx", 4, 1, 7, 36)
write_excel_data("./jobs/21软件7班早午晚检表9.xls", dates)


